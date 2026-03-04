# parsers_core/utils.py
#  ver 4.0  30/01/2026
#  вариант ./parser_core/util.py
#  для main.py 3.5  Переделан write_excel (теперь он дописывает а не перезаписывает )
#  добавлен Общий wrapper для Selenium, и общий блок защиты на случай отсутсвия параметров в config.py

import os
import re
import unicodedata
import json
import pandas as pd
import gc
import openpyxl
from datetime import datetime
import requests
#from openpyxl import load_workbook
import psycopg2
import sqlite3
from typing import Optional, Union
from selenium.common.exceptions import NoSuchWindowException, WebDriverException
import time

# Импортируем config_loader ОТНОСИТЕЛЬНО
try:
    from .config_loader import cfg
except ImportError:
    # Fallback для обратной совместимости
    try:
        from config_loader import cfg
    except ImportError:
        cfg = None
        print("⚠️ config_loader не найден")
"""
# Пробуем получить emoji_agent из конфига
if cfg:
    try:
        emj_agent = cfg.get("", "emoji_agent", "🥷")
    except:
        emj_agent = "🥷"
else:
    # Fallback если config_loader не загружен
    try:
        from config import emoji_agent
        emj_agent = emoji_agent
    except ImportError:
        emj_agent = "🥷"
"""
try:
    # Пробуем импортировать из main
    from main import EMOJI_AGENT
    emj_agent =  EMOJI_AGENT
except ImportError:
    # Запасной вариант
    emj_agent = "🥷"

class SeleniumSoftFail(Exception):
    """Мягкая ошибка Selenium (антибот / закрыто окно / CF)"""
    pass


# =========================
# CONSOLE COLORS (ANSI)
# =========================

ANSI_RESET     = "\x1b[0m"

ANSI_GREEN     = "\x1b[32m"
ANSI_YELLOW    = "\x1b[33m"
ANSI_GRAY      = "\x1b[90m"
ANSI_RED       = "\x1b[31m"
ANSI_PURPLE    = "\x1b[35m"
ANSI_BLACK_BG  = "\x1b[40m"


def set_console_mode(color: str) -> None:
    """
    Устанавливает режим вывода консоли:
    цвет текста + чёрный фон.
    Работает как 'переключатель режима', а не разовая окраска.
    """
    print(color + ANSI_BLACK_BG, end="", flush=True)


def reset_console_mode() -> None:
    """
    Сброс ANSI-режима (возврат к дефолту).
    """
    print(ANSI_RESET, end="", flush=True)

def plural_ru(n: int, one: str, few: str, many: str) -> str:
    """
    Русская форма множественного числа.
    n    — число
    one  — 1 (парсер)
    few  — 2-4 (парсера)
    many — 0, 5+ (парсеров)
    """
    n = abs(int(n))

    if 11 <= n % 100 <= 14:
        return many

    last = n % 10
    if last == 1:
        return one
    if 2 <= last <= 4:
        return few
    return many

def safe_open(sb, url: str, tag: str = "", retries: int = 1, sleep_sec: int = 3) -> bool:
    """
    Безопасное открытие страницы через SeleniumBase.
    НЕ падает.
    Возвращает True / False.
    """

    for attempt in range(1, retries + 1):
        try:
            sb.open(url)
            return True

        except NoSuchWindowException as e:
            _safe_quit(sb)
            if attempt >= retries:
                raise SeleniumSoftFail(
                    f"{tag}: окно браузера было закрыто (CF / anti-bot)"
                ) from e
            time.sleep(sleep_sec)

        except WebDriverException as e:
            _safe_quit(sb)
            if attempt >= retries:
                raise SeleniumSoftFail(
                    f"{tag}: WebDriverException ({e.__class__.__name__})"
                ) from e
            time.sleep(sleep_sec)

        except Exception as e:
            _safe_quit(sb)
            if attempt >= retries:
                raise SeleniumSoftFail(
                    f"{tag}: неизвестная ошибка Selenium: {e}"
                ) from e
            time.sleep(sleep_sec)

    return False

#import json

def safe_json_resp(resp, where=""):
    # curl_cffi Response: resp.text обычно есть
    text = getattr(resp, "text", "") or ""
    ct = ""
    try:
        ct = (resp.headers.get("content-type") or "").lower()
    except Exception:
        pass

    # если вообще пусто
    if not text.strip():
        raise RuntimeError(f"{where} empty body | HTTP {resp.status_code} | ct={ct}")

    # если похоже на HTML/текст, а не JSON
    if ("application/json" not in ct) and (not text.lstrip().startswith(("{", "["))):
        head = text[:200].replace("\n", " ")
        raise RuntimeError(f"{where} not-json | HTTP {resp.status_code} | ct={ct} | head={head!r}")

    try:
        return resp.json()
    except Exception as e:
        head = text[:200].replace("\n", " ")
        raise RuntimeError(f"{where} json-parse-failed | HTTP {resp.status_code} | ct={ct} | head={head!r} | err={e}")

def safe_get_cookies(sb, url: str, tag: str = ""):
    """
    Универсальный безопасный метод получения cookies.
    """
    ok = safe_open(sb, url, tag=tag, retries=1)
    if not ok:
        return {}

    try:
        cookies = sb.driver.get_cookies()
        return {c["name"]: c["value"] for c in cookies}
    except Exception as e:
        raise SeleniumSoftFail(f"{tag}: не удалось получить cookies: {e}") from e


def _safe_quit(sb):
    """
    Гарантированное закрытие драйвера без исключений.
    """
    try:
        sb.quit()
    except Exception:
        pass

def count_records(
    source_path: str,
    *,
    table: Optional[str] = None,
    sheet: Optional[Union[int, str]] = None,
    where: Optional[str] = None
) -> int:
    """
    Универсально считает количество записей в источнике данных.

    Поддерживаемые источники:
    - Excel (.xlsx, .xls)
    - SQLite (.sqlite, .db)
    - PostgreSQL (DSN / connection string)

    Никаких side-effect'ов. Только чтение.

    :param source_path: путь к файлу или строка подключения
    :param table: имя таблицы (для БД)
    :param sheet: имя листа (для Excel)
    :param where: SQL-условие WHERE (без слова WHERE)
    :return: количество записей
    """

    source_path = source_path.strip()

    # =====================================================
    # 1. EXCEL
    # =====================================================
    if source_path.lower().endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(source_path, read_only=True)
        try:
            if sheet is None:
                ws = wb[wb.sheetnames[0]]
            elif isinstance(sheet, int):
                ws = wb[wb.sheetnames[sheet]]
            else:
                ws = wb[sheet]
            return max(ws.max_row - 1, 0)
        finally:
            wb.close()

    # =====================================================
    # 2. SQLITE
    # =====================================================
    if source_path.lower().endswith((".sqlite", ".db")):
        if not os.path.isfile(source_path):
            raise FileNotFoundError(f"SQLite файл не найден: {source_path}")

        if not table:
            raise ValueError("Для SQLite необходимо указать table")

        conn = sqlite3.connect(source_path)
        try:
            cur = conn.cursor()

            sql = f"SELECT COUNT(*) FROM {table}"
            if where:
                sql += f" WHERE {where}"

            cur.execute(sql)
            count = cur.fetchone()[0]
            return int(count)

        finally:
            conn.close()

    # =====================================================
    # 3. POSTGRESQL
    # =====================================================
    # Всё, что не файл — считаем PostgreSQL
    if not table:
        raise ValueError("Для PostgreSQL необходимо указать table")

    conn = psycopg2.connect(source_path)
    try:
        cur = conn.cursor()

        sql = f"SELECT COUNT(*) FROM {table}"
        if where:
            sql += f" WHERE {where}"

        cur.execute(sql)
        count = cur.fetchone()[0]
        return int(count)

    finally:
        conn.close()

# === Функции записи Excel и JSON ===
#====================================================================

def add_metadata_sheet(writer, data, debug_mode, assoc_debug, df):
    """Добавляет лист с метаданными"""
    try:
        # Получаем версию из глобальных переменных
        version = 'N/A'
        try:
            import __main__
            if hasattr(__main__, '__version__'):
                version = __main__.__version__
            elif '__version__' in globals():
                version = globals()['__version__']
        except:
            pass

        # Определяем имя парсера
        parser_name = 'Unknown'
        if data and len(data) > 0:
            parser_name = data[0].get('Сеть', 'Unknown')

        # Создаем DataFrame с метаданными
        metadata = {
            'Параметр': [
                'Режим записи',
                'Время записи',
                'Количество записей',
                'Количество колонок',
                'Парсер',
                'Версия системы',
                'DEBUG_Mode',
                'ASSOC_Debug',
                'Категоризировано',
                'Брендов заполнено',
                'Весов заполнено',
                'Уникальных категорий',
                'Всего строк в файле'
            ],
            'Значение': [
                'DEBUG' if debug_mode else ('ASSOC_DEBUG' if assoc_debug else 'PRODUCTION'),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                len(df),
                len(df.columns),
                parser_name,
                version,
                str(debug_mode),
                str(assoc_debug),
                f"{df['Категория'].notna().sum()}/{len(df)}" if 'Категория' in df.columns else 'N/A',
                f"{df['Бренд'].notna().sum()}/{len(df)}" if 'Бренд' in df.columns else 'N/A',
                f"{df['Вес'].notna().sum()}/{len(df)}" if 'Вес' in df.columns else 'N/A',
                f"{df['Категория'].nunique()}" if 'Категория' in df.columns else 'N/A',
                len(df)
            ]
        }
        meta_df = pd.DataFrame(metadata)
        meta_df.to_excel(writer, sheet_name='Метаданные', index=False)
        print(f"📋 Добавлен лист 'Метаданные'")

    except Exception as meta_error:
        print(f"⚠️ Не удалось добавить метаданные: {meta_error}")
def write_json(data, filename):
    """
    Записывает данные в JSON-файл.
    """
    try:
        folder = os.path.dirname(filename)
        if folder:
            os.makedirs(folder, exist_ok=True)

        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"💾 JSON сохранён: {filename}")
    except Exception as e:
        print(f"❌ Ошибка записи JSON: {e}")


# === Функция логирования ===

def write_log(text, filename="parser_log.txt"):
    """
    Пишет текст в лог-файл с меткой времени.
    """
    try:
        os.makedirs("logs", exist_ok=True)
        filepath = os.path.join("logs", filename)
        with open(filepath, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {text}\n")
        print(f"📝 Лог записан: {text}")
    except Exception as e:
        print(f"⚠️ Ошибка записи в лог: {e}")

    # === Унифицированная функция логирования по ТЗ ===

    # def write_log_uni(text, filename=None):
    """
    Универсальная запись лога.
    Формат строки:
    YYYY-MM-DD \t HH:MM \t <text>
    """

def _load_telegram_bot_cfg() -> Optional[dict]:
    """
    Универсально пытается получить TelegramBot:
    1) config.py (основной для main.py)
    2) scheduler.config (для Scheduler)
    """
    # 1) config.py
    try:
#        from config import TelegramBot as TB
        from parsers_core.tlgcfg import TelegramBot as TB
        if isinstance(TB, dict):
            return TB
    except Exception:
        pass

    # 2) scheduler.config
    try:
        # Используем относительный импорт для scheduler_core
        from ..scheduler_core.config_loader import load_scheduler_config
        cfg = load_scheduler_config()
        tb = cfg.get("telegram_bot")
        if isinstance(tb, dict):
            return tb
    except Exception:
        pass

    return None

# === Основная функция отправки уведомлений в Telegram ===
class TelegramBot:
    """
    Совместимость с кодом, который ожидает TelegramBot в utils.

    Реальная логика уже в функциях SendTelegram() / SendTelegramFile(),
    поэтому здесь просто обёртки.
    """

    @staticmethod
    def get(key: str, default=None):
        cfg = _load_telegram_bot_cfg()
        if not cfg:
            return default
        return cfg.get(key, default)

    @staticmethod
    def send(message: str) -> bool:
        return bool(SendTelegram(message))

    @staticmethod
    def send_file(filepath: str, caption: str = "") -> bool:
        return bool(SendTelegramFile(filepath, caption))

# === Дополнительная функция: отправка файла в Telegram ===
def SendTelegramFile(filepath: str, caption: str = ""):
    """
    Отправляет файл (Excel, лог, JSON и т.д.) в Telegram.
    Использует параметры из config.TelegramBot.
    """
    TelegramBot = _load_telegram_bot_cfg()
    if not TelegramBot:
        return False
    tg_cfg = TelegramBot

    send_flag = TelegramBot.get("send", False)
    if not send_flag:
        print("📴 Отправка файлов в Telegram отключена (send=False).")
        return False

    token = TelegramBot.get("TOKEN")
    chat_id = TelegramBot.get("CHAT_ID")
    my_id = TelegramBot.get("MY_ID")

    if not token or not chat_id:
        print("⚠️ Ошибка: TOKEN или CHAT_ID не указаны в config.TelegramBot")
        return False

    if not os.path.exists(filepath):
        print(f"⚠️ Файл не найден: {filepath}")
        return False

    caption = f"{emj_agent} {my_id} {emj_agent}\n{caption}"

    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(filepath, "rb") as file_data:
        files = {"document": file_data}
        data = {"chat_id": chat_id, "caption": caption}

        try:
            response = requests.post(url, data=data, files=files, timeout=30)
            if response.status_code == 200:
                print(f"📤 Файл {os.path.basename(filepath)} отправлен в Telegram.")
                return True
            else:
                print(f"⚠️ Ошибка Telegram API: {response.status_code} {response.text}")
                return False
        except Exception as e:
            print(f"❌ Ошибка при отправке файла в Telegram: {e}")
            return False

def SendTelegram(message: str):
    """
    Отправляет сообщение в Telegram.
    Использует TOKEN, CHAT_ID и send из config.TelegramBot.
    Поддерживает HTML и эмодзи.
    """

    TelegramBot = _load_telegram_bot_cfg()
    if not TelegramBot:
        return False

    tg_cfg = TelegramBot
    send_flag = TelegramBot.get("send", False)
    if not send_flag:
        print("📴 Отправка в Telegram отключена (send=False).")
        return False

    token = TelegramBot.get("TOKEN")
    chat_id = TelegramBot.get("CHAT_ID")
    my_id = TelegramBot.get("MY_ID")
    time_now = datetime.now().strftime("%H:%M:%S")
    if not token or not chat_id:
        print("⚠️ Ошибка: TOKEN или CHAT_ID не указаны в config.TelegramBot")
        return False

    # ✅ Формируем сообщение с нормальными переводами строк
    message = f"{emj_agent} {my_id} {emj_agent}   ⏱️ {time_now}  \n{message}"

    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = {
        "chat_id": chat_id,
        "text": message,
        "parse_mode": "HTML"
    }


    try:
        response = requests.post(url, data=data, timeout=5)
        if response.status_code != 200:
            # Telegram не критичен — просто выходим
            return False
        return True

    except (requests.exceptions.Timeout,
            requests.exceptions.ConnectionError):
        # ❌ Telegram недоступен — ТИХО игнорируем
        return False

    except Exception:
        # ❌ Любая другая ошибка Telegram — тоже молча
        return False


def debug_mode(cfg_param: dict = None) -> bool:
    """
    DEBUG_MODE считается включённым ТОЛЬКО если:
      DEBUG_MODE = true
      DEBUG_MODE = 1
      DEBUG_MODE = yes

    Во всех остальных случаях — False.

    Args:
        cfg_param: словарь конфигурации (опционально)
    """
    if cfg_param is None:
        # Используем глобальный cfg
        if cfg:
            debug_value = cfg.get("", "DEBUG_MODE", "")
        else:
            # Fallback на config.py
            try:
                from config import DEBUG_MODE
                debug_value = DEBUG_MODE
            except:
                debug_value = ""
    else:
        debug_value = cfg_param.get("DEBUG_MODE", "")

    return str(debug_value).lower() in ("true", "1", "yes")


def update_retail_points(parser_name: str, city: str, count: int) -> bool:
    """
    Обновляет количество торговых точек и фиксирует событие в Excel-файле.
    """
    try:
        from config import db_retail_map
    except Exception:
        print("⚠️ update_retail_points: нет db_retail_map в config.py — пропуск")
        return False

    # --- 1. Корректируем путь ---
    filename = db_retail_map
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    # --- 2. Если файла нет — создаём новый ---
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Дата", "Время", "Имя парсера", "Город", "Количество ТТ"])
        wb.save(filename)
    else:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

    updated = False
    today = datetime.now().strftime("%Y-%m-%d")
    now_time = datetime.now().strftime("%H:%M:%S")

    old_count = None
    found = False

    # --- 3. Ищем строку ---
    for row in ws.iter_rows(min_row=2, max_col=5):
        parser_cell = row[2].value
        city_cell = row[3].value

        if (parser_cell and str(parser_cell).strip().lower() == parser_name.lower() and
                city_cell and str(city_cell).strip().lower() == city.lower()):

            old_count = row[4].value
            found = True

            # Проверяем, нужно ли обновлять
            if old_count is None or count > int(old_count):
                row[0].value = today
                row[1].value = now_time
                row[2].value = parser_name
                row[3].value = city
                row[4].value = count
                updated = True
                print(f"➡️ Обновлено: {parser_name} — {city} количество ТТ 🏪  {old_count} → {count} 🆕🏪   ")

            break

    # --- 4. Если строка не найдена — создаём новую ---
    if not found:
        ws.append([today, now_time, parser_name, city, count])
        updated = True
        print(f"➕ Добавлено: {parser_name} — {city} количество ТТ 🏪 = {count} 🆕")

    # --- 5. Если не нашли и не обновились — старый count остался прежним ---
    if not updated and found:
        print(f"ℹ️ Без изменений: {parser_name} — {city} количество ТТ 🏪 {count} <= {old_count} 📇")

    # --- 6. Сохраняем если были изменения ---
    if updated:
        wb.save(filename)

    return updated


#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||

#====== для v 4.20   ======================================================================================
# Предполагаем, что эти функции уже определены где-то в вашем коде
# from config_loader import cfg
# from utils import SendTelegram, count_records, add_metadata_sheet

def write_sql_draft(data, debug_mode=False, assoc_debug=False):
    """
    Записывает данные в PostgreSQL (таблица retail_draft).
    ВСЕГДА дописывает данные к существующим.

    Три режима работы:
    1. debug_mode=True: все поля (полный дебаг data_enricher)
    2. assoc_debug=True: основные + поля категорий для отладки ассоциаций
    3. Оба False: только production поля для КП

    Нормализация данных:
    - Бренд (нормализованный) → Бренд
    - weight_volume → Вес
    - uni_cat → Категория
    - Артикул → GTIN (если GTIN пустой)
    """

    # Используем cfg из config_loader
    if cfg:
        try:
            SQL = cfg.get("SQL", {})
            TelegramBot = cfg.get("TelegramBot", {})
        except:
            SQL = {}
            TelegramBot = {}
    else:
        # Fallback на старый импорт
        try:
            from config import SQL, TelegramBot
        except Exception as e:
            print(f"⚠️🧯 SQL Draft пропущен — нет SQL/TelegramBot в config.py: {e}")
            return "FAILED"

    # --- базовые проверки ---
    if not isinstance(SQL, dict) or not SQL.get("DB_Draft"):
        print("⚠️🧯 SQL Draft пропущен — нет DB_Draft")
        return "FAILED"

    dbtable = SQL.get("TABLE_Draft")
    if not dbtable:
        print("⚠️🧯 SQL Draft пропущен — TABLE_Draft не задан")
        return "FAILED"

    if not data:
        print("⚠️🧯 SQL Draft пропущен — нет данных")
        return "FAILED"

    print(f"\n{'=' * 60}")
    print(f"📝 ПОДГОТОВКА ДАННЫХ ДЛЯ POSTGRESQL")
    print(f"{'=' * 60}")

    # === ПОДГОТОВКА И НОРМАЛИЗАЦИЯ ДАННЫХ ===
    prepared_data = []

    for item in data:
        new_item = item.copy()

        # === НОРМАЛИЗАЦИЯ ДАННЫХ ===
        # 1. Бренд: копируем нормализованный бренд если есть
        normalized_brand = item.get('Бренд (нормализованный)') or item.get('normalized_brand')
        if normalized_brand:
            new_item['Бренд'] = normalized_brand
            # Сохраняем также в отдельное поле для отладки
            new_item['normalized_brand'] = normalized_brand

        # 2. Вес: копируем weight_volume если поле Вес пустое
        weight_volume = item.get('weight_volume')
        current_weight = item.get('Вес') or item.get('weight')
        if weight_volume and not current_weight:
            new_item['Вес'] = weight_volume

        # 3. Категория: копируем uni_cat в Категория
        uni_cat = item.get('uni_cat')
        if uni_cat:
            new_item['Категория'] = uni_cat

        # 4. GTIN: если GTIN пустой, используем Артикул
        current_gtin = item.get('GTIN') or item.get('gtin')
        article = item.get('Артикул') or item.get('article')
        if not current_gtin and article:
            new_item['GTIN'] = article

        prepared_data.append(new_item)

    # === ВЫБОР ПОЛЕЙ В ЗАВИСИМОСТИ ОТ РЕЖИМА ===
    if debug_mode:
        # В debug_mode показываем все поля (полный дебаг data_enricher)
        mode_label = "DEBUG (data_enricher)"
        data_for_sql = prepared_data
        print(f"🔧 Режим: {mode_label} - все технические поля")

        # Показываем статистику нормализации
        normalized_count = sum(
            1 for item in prepared_data if 'Бренд (нормализованный)' in item and item['Бренд (нормализованный)'])
        weight_fixed = sum(1 for item in prepared_data if item.get('weight_volume') and not item.get('Вес'))
        cat_normalized = sum(1 for item in prepared_data if item.get('uni_cat'))
        print(f"📊 Нормализация: {normalized_count} брендов, {weight_fixed} весов, {cat_normalized} категорий")

    elif assoc_debug:
        # В assoc_debug режиме скрываем технические поля data_enricher, но показываем поля ассоциаций
        mode_label = "ASSOC_DEBUG"
        technical_fields_to_hide = [
            '_data_quality', '_enriched_timestamp', 'brand_source',
            'cat_source', 'normalized_name', 'weight_volume', 'packaging',
            'Бренд (нормализованный)', 'normalized_name'
        ]

        filtered_data = []
        for item in prepared_data:
            new_item = {}
            for key, value in item.items():
                if key not in technical_fields_to_hide:
                    new_item[key] = value
            filtered_data.append(new_item)

        data_for_sql = filtered_data
        print(f"🔍 Режим: {mode_label} - поля для отладки ассоциаций")

    else:
        # В production режиме скрываем все технические и отладочные поля
        mode_label = "PRODUCTION"
        technical_fields_to_hide = [
            '_data_quality', '_enriched_timestamp', 'brand_source',
            'cat_source', 'normalized_name', 'weight_volume', 'packaging',
            'code_cat', 'rule_id', 'confidence', 'retail_cat',
            'Бренд (нормализованный)', 'normalized_name', 'normalized_brand',
            'ID правила', 'Уверенность', 'Категория ритейлера'
        ]

        filtered_data = []
        for item in prepared_data:
            new_item = {}
            for key, value in item.items():
                if key not in technical_fields_to_hide:
                    # Сохраняем нормализованные данные
                    new_item[key] = value
            filtered_data.append(new_item)

        data_for_sql = filtered_data
        print(f"📤 Режим: {mode_label} - только production поля для КП")

    print(f"📊 Записей для записи: {len(data_for_sql)}")

    # === ПОДКЛЮЧЕНИЕ К POSTGRESQL ===
    try:
        print(f"\n🔗 Подключение к PostgreSQL...")
        conn = psycopg2.connect(
            host=SQL.get("host"),
            port=SQL.get("port", 5432),
            user=SQL.get("user"),
            password=SQL.get("password"),
            dbname=SQL.get("DB_Draft")
        )
        conn.autocommit = False
        cursor = conn.cursor()
        print(f"✅ Подключение успешно")

    except Exception as e:
        print(f"❌🗄️🔌 Ошибка подключения к PostgreSQL: {e}")
        print(f"🔄 Переход к SQLite...")
        SendTelegram(f"❌🗄️🔌 SQL Draft: ошибка подключения\n{e}")
        return write_sqlite_draft(prepared_data, debug_mode=debug_mode, assoc_debug=assoc_debug)

    DB_TABLE = dbtable
    date_now = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M:%S")
    agent = TelegramBot.get("MY_ID", "unknown")

    try:
        # === ПРОВЕРКА СУЩЕСТВОВАНИЯ ТАБЛИЦЫ ===
        check_table_sql = f"""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = '{DB_TABLE}'
            );
        """
        cursor.execute(check_table_sql)
        table_exists = cursor.fetchone()[0]

        if table_exists:
            # Получаем количество существующих записей
            count_sql = f'SELECT COUNT(*) FROM "{DB_TABLE}"'
            cursor.execute(count_sql)
            old_count = cursor.fetchone()[0]
            print(f"📊 Таблица существует, записей: {old_count}")
        else:
            old_count = 0
            print(f"📊 Таблица не существует, будет создана")

        # === СОЗДАНИЕ ТАБЛИЦЫ ЕСЛИ НЕ СУЩЕСТВУЕТ ===
        if not table_exists:
            # Базовые поля для всех режимов
            create_table_sql = f"""
                CREATE TABLE "{DB_TABLE}" (
                    id SERIAL PRIMARY KEY,
                    data_date TEXT NOT NULL,
                    data_time TEXT NOT NULL,
                    agent TEXT,
                    retailer TEXT,
                    store_type TEXT,
                    address TEXT,
                    brand TEXT,
                    product_name TEXT,
                    price REAL,
                    promo_price REAL,
                    photo_url TEXT,
                    product_url TEXT,
                    rating REAL,
                    volume TEXT,
                    weight TEXT,
                    stock TEXT,
                    category TEXT,
                    gtin TEXT,
                    flag BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            """

            # Добавляем дополнительные поля для debug режимов
            if debug_mode or assoc_debug:
                create_table_sql += """,
                    rule_id TEXT,
                    confidence REAL,
                    retail_category TEXT,
                    normalized_brand TEXT
                """

                # Дополнительные поля для полного debug_mode
                if debug_mode:
                    create_table_sql += """,
                        code_cat INTEGER,
                        brand_source TEXT,
                        cat_source TEXT,
                        normalized_name TEXT,
                        weight_volume TEXT,
                        packaging TEXT,
                        data_quality TEXT,
                        enriched_timestamp TEXT
                    """

            create_table_sql += ");"

            cursor.execute(create_table_sql)
            print(f"✅ Таблица создана")

        # === ВСТАВКА ДАННЫХ ===
        print(f"\n💾 Вставка {len(data_for_sql)} записей...")

        inserted_count = 0
        skipped_count = 0

        for row in data_for_sql:
            try:
                # === ПОДГОТОВКА ЗНАЧЕНИЙ ===
                # Основные поля
                brand = row.get("Бренд") or row.get("brand") or ""
                product_name = row.get("Название продукта") or row.get("product_name") or ""
                price = row.get("Цена") or row.get("price") or 0
                promo_price = row.get("Цена по акции") or row.get("promo_price") or 0

                # Категория - используем нормализованную если есть
                category = row.get("Категория") or row.get("category") or row.get("uni_cat") or ""

                # Вес - используем нормализованный если есть
                weight = row.get("Вес") or row.get("weight") or row.get("weight_volume") or ""

                # GTIN - используем Артикул если GTIN пустой
                gtin = row.get("GTIN") or row.get("gtin") or ""
                if not gtin:
                    gtin = row.get("Артикул") or row.get("article") or ""

                # Базовые поля и значения
                base_fields = [
                    "data_date", "data_time", "agent", "retailer", "store_type", "address",
                    "brand", "product_name", "price", "promo_price", "photo_url", "product_url",
                    "rating", "volume", "weight", "stock", "category", "gtin", "flag"
                ]

                base_values = [
                    date_now,  # data_date
                    time_now,  # data_time
                    agent,  # agent
                    row.get("Сеть") or row.get("retailer") or "",
                    row.get("Тип магазина") or row.get("store_type") or "",
                    row.get("Адрес Торговой точки") or row.get("address") or "",
                    brand,  # brand (уже нормализованный)
                    product_name,  # product_name
                    price,  # price
                    promo_price,  # promo_price
                    row.get("Фото товара") or row.get("photo_url") or "",
                    row.get("Ссылка на страницу") or row.get("product_url") or "",
                    row.get("Рейтинг") or row.get("rating") or 0,
                    row.get("Объем") or row.get("volume") or "",
                    weight,  # weight (уже нормализованный)
                    row.get("Остаток") or row.get("stock") or "",
                    category,  # category (уже нормализованная)
                    gtin,  # gtin (с Артикула если нужно)
                    False  # flag
                ]

                # Дополнительные поля и значения для debug режимов
                extra_fields = []
                extra_values = []

                if debug_mode or assoc_debug:
                    # Поля для отладки ассоциаций
                    rule_id = row.get("rule_id") or row.get("ID правила") or ""
                    confidence_val = row.get("confidence") or row.get("Уверенность") or 0.0
                    retail_category = row.get("retail_cat") or row.get("Категория ритейлера") or ""
                    normalized_brand_val = row.get("normalized_brand") or brand

                    # Фиксим confidence
                    try:
                        confidence_val = float(confidence_val)
                        if confidence_val > 1.0:
                            confidence_val = 1.0
                        elif confidence_val < 0.0:
                            confidence_val = 0.0
                    except:
                        confidence_val = 0.0

                    extra_fields.extend(["rule_id", "confidence", "retail_category", "normalized_brand"])
                    extra_values.extend([rule_id, confidence_val, retail_category, normalized_brand_val])

                    # Дополнительные поля для полного debug_mode
                    if debug_mode:
                        extra_fields.extend([
                            "code_cat", "brand_source", "cat_source",
                            "normalized_name", "weight_volume", "packaging",
                            "data_quality", "enriched_timestamp"
                        ])

                        extra_values.extend([
                            row.get("code_cat") or -1,
                            row.get("brand_source") or "",
                            row.get("cat_source") or "",
                            row.get("normalized_name") or "",
                            row.get("weight_volume") or "",
                            row.get("packaging") or "",
                            row.get("_data_quality") or "",
                            row.get("_enriched_timestamp") or ""
                        ])

                # Формируем полный запрос
                all_fields = base_fields + extra_fields
                all_values = base_values + extra_values

                placeholders = ", ".join(["%s"] * len(all_fields))
                fields_str = ", ".join(all_fields)

                insert_sql = f"""
                    INSERT INTO "{DB_TABLE}" ({fields_str})
                    VALUES ({placeholders})
                """

                cursor.execute(insert_sql, all_values)
                inserted_count += 1

                # Прогресс каждые 50 записей
                if inserted_count % 50 == 0:
                    print(f"   ... записано {inserted_count} из {len(data_for_sql)}")

            except Exception as row_error:
                skipped_count += 1
                if skipped_count <= 3:  # Показываем только первые 3 ошибки
                    print(f"⚠️ Пропущена запись: {row_error}")

        # === ФИНАЛИЗАЦИЯ ===
        conn.commit()

        # Получаем общее количество записей
        cursor.execute(f'SELECT COUNT(*) FROM "{DB_TABLE}"')
        total_count = cursor.fetchone()[0]
        new_added = total_count - old_count

        print(f"\n✅ POSTGRESQL ({mode_label}): УСПЕШНО")
        print(f"📊 Статистика:")
        print(f"   • Было записей: {old_count}")
        print(f"   • Добавлено новых: {new_added}")
        print(f"   • Всего записей: {total_count}")
        print(f"   • Успешно вставлено: {inserted_count}")
        print(f"   • Пропущено: {skipped_count}")

        # Статистика нормализации
        print(f"\n📊 НОРМАЛИЗАЦИЯ ДАННЫХ:")
        if new_added > 0:
            # Примерная статистика
            print(f"   • Бренды нормализованы: да (если данные были)")
            print(f"   • Веса из weight_volume: да (если нужно было)")
            print(f"   • Категории из uni_cat: да (если были)")
            print(f"   • GTIN из Артикула: да (если GTIN пустой)")

        SendTelegram(f"✅🗄️♻️ PostgreSQL ({mode_label}): +{new_added} записей, всего {total_count}")
        return "POSTGRES"

    except Exception as e:
        # --- ЛЮБАЯ ошибка INSERT → SQLite ---
        conn.rollback()
        print(f"\n❌️🗄️ Ошибка записи в PostgreSQL: {e}")
        print(f"🔄 Переход к SQLite...")
        SendTelegram(f"❌🗄️ SQL Draft: ошибка записи\n{e}")

        try:
            cursor.close()
            conn.close()
        except:
            pass

        return write_sqlite_draft(prepared_data, debug_mode=debug_mode, assoc_debug=assoc_debug)

    finally:
        try:
            cursor.close()
            conn.close()
            print(f"🔗 Соединение закрыто")
            print(f"{'=' * 60}")
        except:
            pass


def write_sqlite_draft(data, debug_mode=False, assoc_debug=False):
    """
    Fallback запись в SQLite при ошибке PostgreSQL
    ВСЕГДА дописывает данные к существующим.

    Идентичная логика нормализации и режимов как в write_sql_draft
    """

    # Используем cfg из config_loader
    if cfg:
        try:
            sqlite_file = cfg.get("", "sqlite_file", "")
            TelegramBot = cfg.get("TelegramBot", {})
            SQL = cfg.get("SQL", {})
        except:
            sqlite_file = ""
            TelegramBot = {}
            SQL = {}
    else:
        # Fallback на старый импорт
        try:
            from config import sqlite_file, TelegramBot, SQL
        except Exception as e:
            print(f"⚠️🛟 SQLite Draft пропущен — нет sqlite_file/SQL/TelegramBot in config.py: {e}")
            return "FAILED"

    DB_TABLE = SQL.get("TABLE_Draft")
    if not DB_TABLE:
        print("⚠️🛟 SQLite Draft пропущен — TABLE_Draft не задан")
        return "FAILED"

    # --- guard: SQLite отключён ---
    if not sqlite_file:
        print("⚠️🛟 SQLite Draft пропущен — sqlite_file не задан в config.py")
        return "FAILED"

    if not data:
        print("⚠️🛟 SQLite Draft пропущен — нет данных")
        return "FAILED"

    print(f"\n{'=' * 60}")
    print(f"📝 ПОДГОТОВКА ДАННЫХ ДЛЯ SQLITE")
    print(f"{'=' * 60}")

    # === ПОДГОТОВКА И НОРМАЛИЗАЦИЯ ДАННЫХ ===
    prepared_data = []

    for item in data:
        new_item = item.copy()

        # Нормализация данных (такая же логика как в PostgreSQL)
        # 1. Бренд
        normalized_brand = item.get('Бренд (нормализованный)') or item.get('normalized_brand')
        if normalized_brand:
            new_item['Бренд'] = normalized_brand
            new_item['normalized_brand'] = normalized_brand

        # 2. Вес
        weight_volume = item.get('weight_volume')
        current_weight = item.get('Вес') or item.get('weight')
        if weight_volume and not current_weight:
            new_item['Вес'] = weight_volume

        # 3. Категория
        uni_cat = item.get('uni_cat')
        if uni_cat:
            new_item['Категория'] = uni_cat

        # 4. GTIN из Артикула
        current_gtin = item.get('GTIN') or item.get('gtin')
        article = item.get('Артикул') or item.get('article')
        if not current_gtin and article:
            new_item['GTIN'] = article

        prepared_data.append(new_item)

    # === ВЫБОР ПОЛЕЙ В ЗАВИСИМОСТИ ОТ РЕЖИМА ===
    if debug_mode:
        mode_label = "DEBUG (data_enricher)"
        data_for_sqlite = prepared_data
        print(f"🔧 Режим: {mode_label}")

    elif assoc_debug:
        mode_label = "ASSOC_DEBUG"
        technical_fields_to_hide = [
            '_data_quality', '_enriched_timestamp', 'brand_source',
            'cat_source', 'normalized_name', 'weight_volume', 'packaging',
            'Бренд (нормализованный)', 'normalized_name'
        ]

        filtered_data = []
        for item in prepared_data:
            new_item = {}
            for key, value in item.items():
                if key not in technical_fields_to_hide:
                    new_item[key] = value
            filtered_data.append(new_item)

        data_for_sqlite = filtered_data
        print(f"🔍 Режим: {mode_label}")

    else:
        mode_label = "PRODUCTION"
        technical_fields_to_hide = [
            '_data_quality', '_enriched_timestamp', 'brand_source',
            'cat_source', 'normalized_name', 'weight_volume', 'packaging',
            'code_cat', 'rule_id', 'confidence', 'retail_cat',
            'Бренд (нормализованный)', 'normalized_name', 'normalized_brand',
            'ID правила', 'Уверенность', 'Категория ритейлера'
        ]

        filtered_data = []
        for item in prepared_data:
            new_item = {}
            for key, value in item.items():
                if key not in technical_fields_to_hide:
                    new_item[key] = value
            filtered_data.append(new_item)

        data_for_sqlite = filtered_data
        print(f"📤 Режим: {mode_label}")

    print(f"📊 Записей для записи: {len(data_for_sqlite)}")

    path = sqlite_file

    # --- создаём директорию при необходимости ---
    try:
        folder = os.path.dirname(path)
        if folder:
            os.makedirs(folder, exist_ok=True)
            print(f"📁 Создана папка: {folder}")
    except Exception:
        pass

    try:
        print(f"\n🔗 Подключение к SQLite...")
        conn = sqlite3.connect(path)
        cur = conn.cursor()
        print(f"✅ Подключение успешно")

        # === ПРОВЕРКА СУЩЕСТВОВАНИЯ ТАБЛИЦЫ ===
        cur.execute(f"""
            SELECT name FROM sqlite_master 
            WHERE type='table' AND name='{DB_TABLE}'
        """)
        table_exists = cur.fetchone() is not None

        if table_exists:
            # Получаем количество существующих записей
            cur.execute(f'SELECT COUNT(*) FROM "{DB_TABLE}"')
            old_count = cur.fetchone()[0]
            print(f"📊 Таблица существует, записей: {old_count}")
        else:
            old_count = 0
            print(f"📊 Таблица не существует, будет создана")

        # --- СОЗДАНИЕ ТАБЛИЦЫ С УЧЕТОМ РЕЖИМА ---
        # Базовые поля
        table_definition = f"""
            CREATE TABLE IF NOT EXISTS "{DB_TABLE}" (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_date TEXT NOT NULL,
                data_time TEXT NOT NULL,
                agent TEXT,
                retailer TEXT,
                store_type TEXT,
                address TEXT,
                brand TEXT,
                product_name TEXT,
                price REAL,
                promo_price REAL,
                photo_url TEXT,
                product_url TEXT,
                rating REAL,
                volume TEXT,
                weight TEXT,
                stock TEXT,
                category TEXT,
                gtin TEXT,
                flag INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        """

        # Добавляем дополнительные поля для debug режимов
        if debug_mode or assoc_debug:
            table_definition += """,
                rule_id TEXT,
                confidence REAL,
                retail_category TEXT,
                normalized_brand TEXT
            """

            # Дополнительные поля для полного debug_mode
            if debug_mode:
                table_definition += """,
                    code_cat INTEGER,
                    brand_source TEXT,
                    cat_source TEXT,
                    normalized_name TEXT,
                    weight_volume TEXT,
                    packaging TEXT,
                    data_quality TEXT,
                    enriched_timestamp TEXT
                """

        table_definition += ")"

        cur.execute(table_definition)

        if not table_exists:
            print(f"✅ Таблица создана")

        date_now = datetime.now().strftime("%Y-%m-%d")
        time_now = datetime.now().strftime("%H:%M:%S")
        agent = TelegramBot.get("MY_ID", "unknown")

        # === ВСТАВКА ДАННЫХ ===
        print(f"\n💾 Вставка {len(data_for_sqlite)} записей...")

        added = 0
        skipped = 0

        for row in data_for_sqlite:
            try:
                # === ПОДГОТОВКА ЗНАЧЕНИЙ ===
                # Основные поля
                brand = row.get("Бренд") or row.get("brand") or ""
                product_name = row.get("Название продукта") or row.get("product_name") or ""

                # Категория
                category = row.get("Категория") or row.get("category") or row.get("uni_cat") or ""

                # Вес
                weight = row.get("Вес") or row.get("weight") or row.get("weight_volume") or ""

                # GTIN
                gtin = row.get("GTIN") or row.get("gtin") or ""
                if not gtin:
                    gtin = row.get("Артикул") or row.get("article") or ""

                # Базовые значения
                base_values = [
                    date_now,  # data_date
                    time_now,  # data_time
                    agent,  # agent
                    row.get("Сеть") or row.get("retailer") or "",
                    row.get("Тип магазина") or row.get("store_type") or "",
                    row.get("Адрес Торговой точки") or row.get("address") or "",
                    brand,  # brand
                    product_name,  # product_name
                    row.get("Цена") or row.get("price") or 0,
                    row.get("Цена по акции") or row.get("promo_price") or 0,
                    row.get("Фото товара") or row.get("photo_url") or "",
                    row.get("Ссылка на страницу") or row.get("product_url") or "",
                    row.get("Рейтинг") or row.get("rating") or 0,
                    row.get("Объем") or row.get("volume") or "",
                    weight,  # weight
                    row.get("Остаток") or row.get("stock") or "",
                    category,  # category
                    gtin,  # gtin
                    0  # flag
                ]

                # Дополнительные значения для debug режимов
                extra_values = []
                if debug_mode or assoc_debug:
                    rule_id = row.get("rule_id") or row.get("ID правила") or ""
                    confidence_val = row.get("confidence") or row.get("Уверенность") or 0.0
                    retail_category = row.get("retail_cat") or row.get("Категория ритейлера") or ""
                    normalized_brand_val = row.get("normalized_brand") or brand

                    # Фиксим confidence
                    try:
                        confidence_val = float(confidence_val)
                        if confidence_val > 1.0:
                            confidence_val = 1.0
                        elif confidence_val < 0.0:
                            confidence_val = 0.0
                    except:
                        confidence_val = 0.0

                    extra_values.extend([rule_id, confidence_val, retail_category, normalized_brand_val])

                    # Дополнительные поля для полного debug_mode
                    if debug_mode:
                        extra_values.extend([
                            row.get("code_cat") or -1,
                            row.get("brand_source") or "",
                            row.get("cat_source") or "",
                            row.get("normalized_name") or "",
                            row.get("weight_volume") or "",
                            row.get("packaging") or "",
                            row.get("_data_quality") or "",
                            row.get("_enriched_timestamp") or ""
                        ])

                # Формируем полный список значений
                all_values = base_values + extra_values

                # Формируем SQL запрос
                placeholders = ", ".join(["?"] * len(all_values))
                cur.execute(f"""
                    INSERT INTO "{DB_TABLE}" VALUES (NULL, {placeholders})
                """, all_values)

                added += 1

                # Прогресс каждые 50 записей
                if added % 50 == 0:
                    print(f"   ... записано {added} из {len(data_for_sqlite)}")

            except Exception as row_error:
                skipped += 1
                if skipped <= 3:  # Показываем только первые 3 ошибки
                    print(f"⚠️ Пропущена запись: {row_error}")

        # === ФИНАЛИЗАЦИЯ ===
        conn.commit()
        conn.close()

        # Получаем общее количество записей
        n = count_records(sqlite_file, table=DB_TABLE)
        new_added = n - old_count

        print(f"\n✅ SQLITE ({mode_label}): УСПЕШНО")
        print(f"📊 Статистика:")
        print(f"   • Было записей: {old_count}")
        print(f"   • Добавлено новых: {new_added}")
        print(f"   • Всего записей: {n}")
        print(f"   • Успешно вставлено: {added}")
        print(f"   • Пропущено: {skipped}")
        print(f"   📁 Файл: {path}")

        # Статистика нормализации
        if new_added > 0:
            print(f"\n📊 НОРМАЛИЗАЦИЯ ДАННЫХ:")
            print(f"   • Данные нормализованы: да (бренд, вес, категория, GTIN)")

        SendTelegram(f"💾🛟 SQLite ({mode_label}): +{new_added} записей, всего {n}")
        return "SQLITE"

    except Exception as e:
        print(f"\n❌️🛟 SQLite Draft: ошибка записи\n{e}")
        SendTelegram(f"❌️🛟 SQLite Draft: ошибка записи\n{e}")
        try:
            conn.close()
        except:
            pass
        print(f"{'=' * 60}")
        return "FAILED"

#====================================================================================================
def write_excel(data, filename, sheet_name="Data", debug_mode=False, assoc_debug=False):
    """
    Записывает данные в Excel файл
    Всегда дописывает данные к существующему файлу

    Три режима работы:
    1. debug_mode=True: все поля (полный дебаг data_enricher)
    2. assoc_debug=True: основные + поля категорий для отладки ассоциаций
    3. Оба False: только production поля для КП

    Нормализация данных:
    - Бренд (нормализованный) → Бренд
    - weight_volume → Вес
    - uni_cat → Категория
    - Артикул → GTIN (если GTIN пустой)
    """
    if not data:
        print("⚠️ Нет данных для записи в Excel")
        return

    try:
        # Конвертируем в DataFrame
        df = pd.DataFrame(data)

        print(f"\n{'=' * 60}")
        print(f"📝 ПОДГОТОВКА ДАННЫХ ДЛЯ EXCEL")
        print(f"{'=' * 60}")

        # === НОРМАЛИЗАЦИЯ ДАННЫХ В DATAFRAME ===

        # 1. Бренд: копируем нормализованный бренд если есть
        if 'Бренд (нормализованный)' in df.columns:
            # Заменяем пустые или NaN значения в основном Бренде нормализованными
            mask = df['Бренд (нормализованный)'].notna()
            if mask.any():
                df.loc[mask, 'Бренд'] = df.loc[mask, 'Бренд (нормализованный)']
                normalized_count = mask.sum()
                print(f"🏷️  Нормализовано {normalized_count} брендов")

        # 2. Вес: копируем weight_volume если поле Вес пустое
        if 'weight_volume' in df.columns:
            # Создаем колонку 'Вес', если она отсутствует
            if 'Вес' not in df.columns:
                df['Вес'] = None
            mask = df['weight_volume'].notna() & (df['Вес'].isna() | (df['Вес'] == ''))
            if mask.any():
                df.loc[mask, 'Вес'] = df.loc[mask, 'weight_volume']
                weight_fixed = mask.sum()
                print(f"⚖️  Перенесено {weight_fixed} весов/объемов из weight_volume")

        # 3. Категория: копируем uni_cat в Категория
        if 'uni_cat' in df.columns:
            # Создаем колонку Категория если нет
            if 'Категория' not in df.columns:
                df['Категория'] = df['uni_cat']
            else:
                # Заполняем пустые категории из uni_cat
                mask = df['uni_cat'].notna() & (df['Категория'].isna() | (df['Категория'] == ''))
                if mask.any():
                    df.loc[mask, 'Категория'] = df.loc[mask, 'uni_cat']
            cat_normalized = df['uni_cat'].notna().sum()
            print(f"📦  Нормализовано {cat_normalized} категорий из uni_cat")

        # 4. GTIN: если GTIN пустой, используем Артикул
        if 'Артикул' in df.columns:
            # Проверяем есть ли колонка GTIN
            if 'GTIN' not in df.columns:
                df['GTIN'] = df['Артикул']
            else:
                # Заполняем пустые GTIN из Артикула
                mask = df['Артикул'].notna() & (df['GTIN'].isna() | (df['GTIN'] == ''))
                if mask.any():
                    df.loc[mask, 'GTIN'] = df.loc[mask, 'Артикул']
                    gtin_fixed = mask.sum()
                    print(f"🔢  Заполнено {gtin_fixed} GTIN из Артикулов")

        # === ВЫБОР ПОЛЕЙ В ЗАВИСИМОСТИ ОТ РЕЖИМА ===

        # Определяем режим
        if debug_mode:
            mode_name = "DEBUG (data_enricher)"
            print(f"🔧 Режим: {mode_name} - все технические поля")
        elif assoc_debug:
            mode_name = "ASSOC_DEBUG"
            print(f"🔍 Режим: {mode_name} - поля для отладки ассоциаций")
        else:
            mode_name = "PRODUCTION"
            print(f"📤 Режим: {mode_name} - только production поля для КП")

        # Списки полей для разных режимов
        # Основные поля (всегда показываем в production)
        main_fields = [
            'Номер', 'Сеть', 'Тип магазина', 'Адрес Торговой точки',
            'Бренд', 'Название продукта', 'Цена', 'Цена по акции',
            'Фото товара', 'Ссылка на страницу', 'Рейтинг',
            'Объем', 'Вес', 'Остаток', 'Категория', 'GTIN'
        ]

        # Поля категорий (только для assoc_debug и debug)
        category_fields = ['ID правила', 'Уверенность', 'Категория ритейлера']

        # Технические поля data_enricher (только для debug_mode)
        technical_fields = [
            'uni_cat', 'code_cat', 'brand', 'brand_source',
            'normalized_name', 'weight_volume', 'packaging',
            '_data_quality', '_enriched_timestamp', 'cat_source',
            'Бренд (нормализованный)', 'rule_id', 'confidence', 'retail_cat'
        ]

        # Формируем окончательный список полей
        if debug_mode:
            # Все поля включая технические
            all_fields = main_fields + technical_fields
            # Добавляем все что еще есть
            other_fields = [f for f in df.columns if f not in all_fields]
            all_fields = all_fields + other_fields

        elif assoc_debug:
            # Основные + поля категорий, без технических полей data_enricher
            all_fields = main_fields + category_fields
            # Убираем технические поля
            all_fields = [f for f in all_fields if f not in technical_fields]
            # Добавляем все что еще есть (кроме технических)
            other_fields = [f for f in df.columns if f not in all_fields and f not in technical_fields]
            all_fields = all_fields + other_fields

        else:
            # PRODUCTION: ТОЛЬКО основные поля
            all_fields = main_fields.copy()
            # Убираем все технические поля
            technical_fields_to_exclude = [
                'uni_cat', 'code_cat', 'brand', 'brand_source',
                'normalized_name', 'weight_volume', 'packaging',
                '_data_quality', '_enriched_timestamp', 'cat_source',
                'Бренд (нормализованный)', 'rule_id', 'confidence',
                'retail_cat', 'ID правила', 'Уверенность', 'Категория ритейлера'
            ]
            all_fields = [f for f in all_fields if f not in technical_fields_to_exclude]

            # Добавляем только не-технические поля
            other_fields = [f for f in df.columns
                            if f not in all_fields
                            and f not in technical_fields_to_exclude
                            and not f.startswith('_')  # убираем все поля с подчеркиванием
                            and not f.endswith('_source')
                            and not f.endswith('_timestamp')
                            and 'normalized' not in f.lower()]
            all_fields = all_fields + other_fields

        # Оставляем только существующие поля
        existing_fields = [field for field in all_fields if field in df.columns]

        # Убираем дубликаты, сохраняя порядок
        seen = set()
        final_fields = [x for x in existing_fields if not (x in seen or seen.add(x))]

        print(f"📊 Колонки в {mode_name} режиме ({len(final_fields)}):")
        for i, col in enumerate(final_fields[:15], 1):
            print(f"   {i:2d}. {col}")
        if len(final_fields) > 15:
            print(f"   ... и еще {len(final_fields) - 15} колонок")

        # Выбираем только нужные поля
        df = df[final_fields]

        # === ПЕРЕИМЕНОВАНИЕ КОЛОНОК ДЛЯ ЧИТАЕМОСТИ (если нужно) ===
        # В production режиме переименования не нужны, так как технических полей нет

        if debug_mode or assoc_debug:
            column_renames = {}

            if debug_mode:
                debug_renames = {
                    'code_cat': 'Код категории',
                    'brand_source': 'Источник бренда',
                    'cat_source': 'Источник категории',
                    'normalized_name': 'Нормализованное название',
                    'weight_volume': 'Вес/объем (тех.)',
                    'packaging': 'Упаковка',
                    '_data_quality': 'Качество данных',
                    '_enriched_timestamp': 'Время обогащения',
                    'Бренд (нормализованный)': 'Бренд (норм., тех.)',
                    'uni_cat': 'Категория (исходная)',
                    'rule_id': 'ID правила',
                    'confidence': 'Уверенность',
                    'retail_cat': 'Категория ритейлера'
                }

                for old_name, new_name in debug_renames.items():
                    if old_name in df.columns:
                        column_renames[old_name] = new_name

            elif assoc_debug:
                assoc_renames = {
                    'rule_id': 'ID правила',
                    'confidence': 'Уверенность',
                    'retail_cat': 'Категория ритейлера'
                }

                for old_name, new_name in assoc_renames.items():
                    if old_name in df.columns:
                        column_renames[old_name] = new_name

            if column_renames:
                df = df.rename(columns=column_renames)

        # === СТАТИСТИКА ПО КОЛОНКАМ ===
        print(f"\n📊 СТАТИСТИКА ДАННЫХ:")

        # Основные колонки
        essential_cols = ['Бренд', 'Вес', 'Категория', 'GTIN']
        for col in essential_cols:
            if col in df.columns:
                filled = df[col].notna().sum()
                total = len(df)
                percent = (filled / total * 100) if total > 0 else 0
                status = "✓" if percent > 80 else "⚠️"
                print(f"   {status} {col}: {filled}/{total} заполнено ({percent:.1f}%)")

        # Категории
        if 'Категория' in df.columns:
            categorized = df['Категория'].notna().sum()
            total = len(df)
            if total > 0:
                percent = categorized / total * 100
                unique_cats = df['Категория'].nunique()
                status = "✓" if percent > 80 else "⚠️"
                print(f"   {status} Категории: {categorized}/{total} ({percent:.1f}%), уникальных: {unique_cats}")

        # === ЗАПИСЬ В EXCEL (ВСЕГДА ДОПИСЫВАЕТ) ===
        print(f"\n💾 ЗАПИСЬ В EXCEL:")

        file_exists = os.path.exists(filename)

        if file_exists:
            print(f"📁 Файл существует: {filename}")

            try:
                # Загружаем существующие данные
                existing_df = pd.read_excel(filename, sheet_name=sheet_name)
                old_count = len(existing_df)
                print(f"📊 В файле уже есть: {old_count} строк")

                # Проверяем совместимость колонок
                if not existing_df.columns.equals(df.columns):
                    print("⚠️ Колонки не совпадают, пытаемся объединить...")

                    # Объединяем колонки из старого и нового DataFrame
                    all_columns = list(dict.fromkeys(list(existing_df.columns) + list(df.columns)))

                    # Переиндексируем оба DataFrame чтобы иметь одинаковые колонки
                    existing_df = existing_df.reindex(columns=all_columns)
                    df = df.reindex(columns=all_columns)

                    print(f"   Объединенные колонки: {len(all_columns)}")

                # Объединяем старые и новые данные
                combined_df = pd.concat([existing_df, df], ignore_index=True, sort=False)
                new_count = len(combined_df)
                added_count = new_count - old_count

                print(f"📈 Объединяем: {old_count} старых + {len(df)} новых")
                print(f"📊 Итого будет: {new_count} строк (+{added_count})")

                # Сохраняем объединенные данные
                with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Добавляем метаданные для debug режимов
                    if (debug_mode or assoc_debug) and len(data) > 0:
                        add_metadata_sheet(writer, data, debug_mode, assoc_debug, combined_df)

                print(f"✅ Данные успешно дописаны")

            except Exception as e:
                print(f"❌ Ошибка при чтении/объединении файла: {e}")
                print(f"📝 Создаем новый файл с текущими данными")

                # Создаем новый файл с текущими данными
                with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    if (debug_mode or assoc_debug) and len(data) > 0:
                        add_metadata_sheet(writer, data, debug_mode, assoc_debug, df)

                print(f"📁 Создан новый файл")

        else:
            print(f"📁 Файл не существует, создаем новый: {filename}")

            # Создаем новый файл
            with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                if (debug_mode or assoc_debug) and len(data) > 0:
                    add_metadata_sheet(writer, data, debug_mode, assoc_debug, df)

            print(f"📝 Создан новый файл с {len(df)} строками")

        # === ФИНАЛЬНАЯ ПРОВЕРКА ===
        try:
            final_df = pd.read_excel(filename, sheet_name=sheet_name)
            print(f"\n✅ ФАЙЛ УСПЕШНО СОХРАНЕН")
            print(f"📁 Файл: {filename}")
            print(f"📊 Итоговое количество строк: {len(final_df)}")
            print(f"📊 Итоговое количество колонок: {len(final_df.columns)}")

            # Показываем все колонки
            cols_list = final_df.columns.tolist()
            print(f"📋 Колонки в файле ({len(cols_list)}):")
            for i, col in enumerate(cols_list[:20], 1):
                print(f"   {i:2d}. {col}")
            if len(cols_list) > 20:
                print(f"   ... и еще {len(cols_list) - 20} колонок")

        except Exception as e:
            print(f"⚠️ Не удалось проверить сохраненный файл: {e}")

        print(f"{'=' * 60}")

    except Exception as e:
        print(f"\n❌ ОШИБКА ЗАПИСИ В EXCEL: {e}")
        import traceback
        traceback.print_exc()
        print(f"{'=' * 60}")

#====== для v 4.20   ======================================================================================

#||||||| v4.22 compatibility helper  ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
# Safe names (table/file)
# v4.22 compatibility helper

def MakeSafeTableName(name: str, max_len: int = 80) -> str:
    """
    Делает безопасное имя для Excel/SQLite/файлов:
    - заменяет запрещённые символы на '_'
    - схлопывает повторяющиеся '_'
    - обрезает до max_len
    - гарантирует непустое имя
    """
    if name is None:
        name = ""
    name = str(name).strip()

    # заменяем всё кроме букв/цифр/._- на "_"
    safe = re.sub(r"[^0-9A-Za-zА-Яа-я._-]+", "_", name, flags=re.UNICODE)
    safe = re.sub(r"_+", "_", safe).strip("._- ")

    if not safe:
        safe = "table"

    # ограничим длину
    if max_len and len(safe) > max_len:
        safe = safe[:max_len].rstrip("._- ")

    return safe


def MakeSafeFileName(name: str, max_len: int = 120) -> str:
    """
    Алиас под безопасное имя файла (если где-то в коде так зовут).
    """
    return MakeSafeTableName(name, max_len=max_len)
# -----------------------------
# Excel save compatibility
# v4.22 helper (for main.py)
# -----------------------------

def SaveToExcel(*args, **kwargs):
    """
    Совместимость с main.py v4.20–v4.22

    Если в проекте есть реальная функция сохранения Excel
    (например WriteExcel / ExportToExcel / save_xlsx),
    сюда можно прокинуть вызов.

    Пока — безопасная заглушка:
    main.py не упадёт, логика парсеров сохранится.
    """
    try:
        # Попытка найти реальную функцию (если она есть под другим именем)
        if 'WriteExcel' in globals():
            return globals()['WriteExcel'](*args, **kwargs)

        if 'ExportToExcel' in globals():
            return globals()['ExportToExcel'](*args, **kwargs)

        if 'save_xlsx' in globals():
            return globals()['save_xlsx'](*args, **kwargs)

        # Если ни одной нет — просто сообщаем
        print("⚠️ SaveToExcel: реальная функция сохранения не найдена, пропуск.")

    except Exception as e:
        print(f"❌ SaveToExcel error: {e}")

# -----------------------------
# Text normalization
# v4.22 compatibility helper
# -----------------------------
def NormalizeText(text: str) -> str:
    """
    Универсальная нормализация текста:
    - None -> ""
    - приведение к str
    - Unicode NFKD
    - lower()
    - удаление лишних пробелов
    - схлопывание мусорных символов

    Поведение безопасное и детерминированное,
    подходит для сравнения, логов, ассоциаций.
    """
    if text is None:
        return ""

    try:
        s = str(text)
    except Exception:
        return ""

    # Unicode normalize
    s = unicodedata.normalize("NFKD", s)

    # lowercase
    s = s.lower()

    # заменить всё кроме букв/цифр на пробел
    s = re.sub(r"[^0-9a-zа-яё]+", " ", s, flags=re.IGNORECASE)

    # схлопнуть пробелы
    s = re.sub(r"\s+", " ", s).strip()

    return s
# -----------------------------
# Time formatting helper
# v4.22 compatibility
# -----------------------------

def FormatElapsed(seconds: float) -> str:
    """
    Форматирует прошедшее время в читабельный вид.
    Примеры:
      5        -> 00:00:05
      65       -> 00:01:05
      3671     -> 01:01:11
      93784    -> 1д 02:03:04

    Используется для логов, Telegram, статуса.
    """
    try:
        sec = int(seconds)
    except Exception:
        return "00:00:00"

    if sec < 0:
        sec = 0

    days, rem = divmod(sec, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, seconds = divmod(rem, 60)

    if days > 0:
        return f"{days}д {hours:02d}:{minutes:02d}:{seconds:02d}"
    else:
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

# -----------------------------
# ETA formatting helper
# v4.22 compatibility
# -----------------------------

def FormatETA(seconds: float) -> str:
    """
    Форматирует ETA (оставшееся время).
    Правила:
      - None/NaN/ошибка -> "ETA ?"
      - <=0 -> "ETA 00:00:00"
      - >0  -> "ETA 00:MM:SS" / "ETA HH:MM:SS" / "ETA 2д HH:MM:SS"
    """
    try:
        if seconds is None:
            return "ETA ?"
        sec = int(float(seconds))
    except Exception:
        return "ETA ?"

    if sec <= 0:
        return "ETA 00:00:00"

    days, rem = divmod(sec, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, seconds = divmod(rem, 60)

    if days > 0:
        return f"ETA {days}д {hours:02d}:{minutes:02d}:{seconds:02d}"
    else:
        return f"ETA {hours:02d}:{minutes:02d}:{seconds:02d}"
# -----------------------------
# Status bar / progress helper
# v4.22 compatibility
# -----------------------------

def MakeStatusBar(current: int, total: int, width: int = 20) -> str:
    """
    Возвращает текстовый progress-bar.
    Используется для print / Telegram / Status.

    Пример:
      [████████░░░░░░░░░░░░] 40% (4/10)
    """
    try:
        cur = int(current)
        tot = int(total)
    except Exception:
        return "[????????????????????] ?% (0/?)"

    if tot <= 0:
        return "[░" * width + "] 0% (0/0)"

    if cur < 0:
        cur = 0
    if cur > tot:
        cur = tot

    pct = int((cur / tot) * 100)
    filled = int((cur / tot) * width)
    empty = width - filled

    bar = "█" * filled + "░" * empty
    return f"[{bar}] {pct:3d}% ({cur}/{tot})"

def ClearStatusBar() -> None:
    """
    Очищает строку прогресса в консоли/терминале.
    В v4.20 могло использоваться для "сброса" статуса после завершения.
    """
    try:
        # carriage return + очистка строки
        print("\r" + " " * 140 + "\r", end="", flush=True)
    except Exception:
        # если stdout/консоль недоступны — молча игнорируем
        pass

# =============================
# Colorama shim (Fore / Style)
# Совместимость с main.py 4.20–4.22
# =============================

try:
    from colorama import Fore as _Fore, Style as _Style
except Exception:
    # Fallback если colorama не установлена
    class _Fore:
        BLACK = RED = GREEN = YELLOW = BLUE = MAGENTA = CYAN = WHITE = RESET = ""

    class _Style:
        DIM = NORMAL = BRIGHT = RESET_ALL = ""

# Экспортируем ожидаемые имена
Fore = _Fore
Style = _Style

#||||||| v4.22 compatibility helper  ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||

